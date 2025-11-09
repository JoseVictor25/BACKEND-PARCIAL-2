"""
Utilidades para generar reportes en diferentes formatos.
"""
from io import BytesIO
from django.utils import timezone
from django.db.models import Sum, Count, Avg
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


def generar_reporte_ventas_pdf(datos, fecha_inicio=None, fecha_fin=None):
    """
    Genera un reporte de ventas en formato PDF.
    """
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []
    
    # Estilos
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#2c3e50'),
        spaceAfter=30,
        alignment=TA_CENTER
    )
    
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Normal'],
        fontSize=12,
        textColor=colors.HexColor('#7f8c8d'),
        spaceAfter=12,
        alignment=TA_CENTER
    )
    
    # Título
    elements.append(Paragraph("SmartSales365", title_style))
    elements.append(Paragraph("Reporte de Ventas", subtitle_style))
    
    # Período
    if fecha_inicio and fecha_fin:
        periodo = f"Período: {fecha_inicio.strftime('%d/%m/%Y')} - {fecha_fin.strftime('%d/%m/%Y')}"
        elements.append(Paragraph(periodo, subtitle_style))
    
    elements.append(Spacer(1, 20))
    
    # Resumen general
    resumen_data = [
        ['Métrica', 'Valor'],
        ['Total de Ventas', f"${datos['total_ventas']:.2f}"],
        ['Cantidad de Ventas', str(datos['cantidad_ventas'])],
        ['Ticket Promedio', f"${datos['ticket_promedio']:.2f}"],
        ['Productos Vendidos', str(datos['productos_vendidos'])]
    ]
    
    resumen_table = Table(resumen_data, colWidths=[3*inch, 2*inch])
    resumen_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498db')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 14),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(resumen_table)
    elements.append(Spacer(1, 30))
    
    # Tabla de ventas detalladas
    if datos.get('ventas_detalle'):
        elements.append(Paragraph("Detalle de Ventas", styles['Heading2']))
        elements.append(Spacer(1, 12))
        
        ventas_data = [['ID', 'Cliente', 'Fecha', 'Total', 'Estado']]
        for venta in datos['ventas_detalle']:
            ventas_data.append([
                str(venta['id']),
                venta['usuario'],
                venta['fecha'],
                f"${venta['total']:.2f}",
                venta['estado']
            ])
        
        ventas_table = Table(ventas_data, colWidths=[0.7*inch, 2*inch, 1.5*inch, 1.2*inch, 1.2*inch])
        ventas_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#34495e')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.lightgrey),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
        ]))
        
        elements.append(ventas_table)
    
    # Pie de página
    elements.append(Spacer(1, 30))
    fecha_gen = timezone.now().strftime('%d/%m/%Y %H:%M:%S')
    pie = Paragraph(
        f"<i>Reporte generado el {fecha_gen}</i>",
        ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8, textColor=colors.grey, alignment=TA_RIGHT)
    )
    elements.append(pie)
    
    doc.build(elements)
    buffer.seek(0)
    return buffer


def generar_reporte_ventas_excel(datos, fecha_inicio=None, fecha_fin=None):
    """
    Genera un reporte de ventas en formato Excel.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte de Ventas"
    
    # Estilos
    header_fill = PatternFill(start_color="3498db", end_color="3498db", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=16, color="2c3e50")
    
    # Título
    ws['A1'] = "SmartSales365 - Reporte de Ventas"
    ws['A1'].font = title_font
    ws.merge_cells('A1:E1')
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Período
    if fecha_inicio and fecha_fin:
        ws['A2'] = f"Período: {fecha_inicio.strftime('%d/%m/%Y')} - {fecha_fin.strftime('%d/%m/%Y')}"
        ws.merge_cells('A2:E2')
        ws['A2'].alignment = Alignment(horizontal='center')
    
    # Resumen
    row = 4
    ws[f'A{row}'] = "RESUMEN GENERAL"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    ws.merge_cells(f'A{row}:B{row}')
    
    row += 1
    resumen_items = [
        ('Total de Ventas:', f"${datos['total_ventas']:.2f}"),
        ('Cantidad de Ventas:', str(datos['cantidad_ventas'])),
        ('Ticket Promedio:', f"${datos['ticket_promedio']:.2f}"),
        ('Productos Vendidos:', str(datos['productos_vendidos']))
    ]
    
    for label, valor in resumen_items:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = valor
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
    
    # Detalle de ventas
    if datos.get('ventas_detalle'):
        row += 2
        ws[f'A{row}'] = "DETALLE DE VENTAS"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws.merge_cells(f'A{row}:E{row}')
        
        row += 1
        headers = ['ID', 'Cliente', 'Fecha', 'Total', 'Estado']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        row += 1
        for venta in datos['ventas_detalle']:
            ws.cell(row=row, column=1, value=venta['id'])
            ws.cell(row=row, column=2, value=venta['usuario'])
            ws.cell(row=row, column=3, value=venta['fecha'])
            ws.cell(row=row, column=4, value=f"${venta['total']:.2f}")
            ws.cell(row=row, column=5, value=venta['estado'])
            row += 1
    
    # Ajustar anchos de columna
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = max_length + 2
    
    # Guardar en buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def generar_datos_reporte_ventas(fecha_inicio=None, fecha_fin=None):
    """
    Obtiene los datos para el reporte de ventas.
    """
    from venta.models import Venta, DetalleVenta
    from django.db.models import Q
    
    # Filtrar ventas por fecha
    ventas_query = Venta.objects.all()
    if fecha_inicio:
        ventas_query = ventas_query.filter(fecha__gte=fecha_inicio)
    if fecha_fin:
        ventas_query = ventas_query.filter(fecha__lte=fecha_fin)
    
    # Calcular métricas
    total_ventas = ventas_query.filter(estado='pagado').aggregate(
        total=Sum('total')
    )['total'] or 0
    
    cantidad_ventas = ventas_query.filter(estado='pagado').count()
    ticket_promedio = total_ventas / cantidad_ventas if cantidad_ventas > 0 else 0
    
    productos_vendidos = DetalleVenta.objects.filter(
        venta__in=ventas_query.filter(estado='pagado')
    ).aggregate(total=Sum('cantidad'))['total'] or 0
    
    # Detalle de ventas
    ventas_detalle = []
    for venta in ventas_query.order_by('-fecha')[:50]:  # Últimas 50 ventas
        ventas_detalle.append({
            'id': venta.id,
            'usuario': venta.usuario.username,
            'fecha': venta.fecha.strftime('%d/%m/%Y %H:%M'),
            'total': float(venta.total),
            'estado': venta.get_estado_display()
        })
    
    return {
        'total_ventas': float(total_ventas),
        'cantidad_ventas': cantidad_ventas,
        'ticket_promedio': float(ticket_promedio),
        'productos_vendidos': productos_vendidos,
        'ventas_detalle': ventas_detalle
    }